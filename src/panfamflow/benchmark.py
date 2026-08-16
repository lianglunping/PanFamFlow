"""Fail-closed biological benchmark intake and readiness auditing."""

from __future__ import annotations

import csv
import gzip
import hashlib
import html
import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from importlib import resources
from pathlib import Path
from typing import Annotated, Any, Literal

import yaml
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    StringConstraints,
    ValidationError,
    model_validator,
)

BenchmarkId = Annotated[
    str,
    StringConstraints(pattern=r"^[A-Za-z][A-Za-z0-9_.-]*$", min_length=1, max_length=100),
]
ApprovalState = Literal["unresolved", "proposed", "approved", "rejected"]
AssemblyLevel = Literal["chromosome", "scaffold", "contig", "unknown"]
InputKind = Literal["assembled_genome", "reference_aligned_sample"]
ChromosomePolicy = Literal["ignore", "warn", "block"]
CheckSeverity = Literal["BLOCK", "WARN", "INFO"]
CheckStatus = Literal["PASS", "FAIL", "WARN", "INFO"]
OverallStatus = Literal["READY", "REVIEW", "BLOCKED"]

_SHA256_RE = re.compile(r"^[0-9a-fA-F]{64}$")
_PLACEHOLDER_NAMES = {"", "TARGET_FAMILY", "UNRESOLVED", "TODO", "TBD", "XX"}
_REQUIRED_SPECIES_PATHS = ("genome", "gff3", "protein", "cds")


class StrictBenchmarkModel(BaseModel):
    """Base model rejecting unknown benchmark fields."""

    model_config = ConfigDict(extra="forbid", validate_assignment=True)


class BenchmarkProject(StrictBenchmarkModel):
    """Benchmark project identity and deterministic seed."""

    name: BenchmarkId = "rice_target_family_pilot"
    analysis_scope: Literal["target_pan_gene_family"] = "target_pan_gene_family"
    seed: int = 20260807
    description: str | None = None


class BenchmarkFamily(StrictBenchmarkModel):
    """Frozen target-family definition required before a biological run."""

    name: str = "UNRESOLVED"
    approval_state: ApprovalState = "unresolved"
    pfam_ids: list[str] = Field(default_factory=list)
    interpro_ids: list[str] = Field(default_factory=list)
    hmm: Path | None = None
    reference_proteins: Path | None = None
    manual_truth_set: Path = Path("manual_review/manual_truth_set.tsv")
    notes: str | None = None


class BenchmarkPanel(StrictBenchmarkModel):
    """Candidate assembled-genome panel and intended pilot size."""

    species_table: Path = Path("species.tsv")
    min_genomes: int = Field(default=5, ge=2)
    max_genomes: int = Field(default=10, ge=2)

    @model_validator(mode="after")
    def validate_panel_bounds(self) -> BenchmarkPanel:
        if self.max_genomes < self.min_genomes:
            raise ValueError("panel.max_genomes must be >= panel.min_genomes")
        return self


class BenchmarkAcceptance(StrictBenchmarkModel):
    """Pre-registered software and biological intake gates."""

    approval_state: ApprovalState = "unresolved"
    require_family_approved: bool = True
    require_sha256: bool = True
    require_manual_truth_set: bool = True
    min_positive_controls: int = Field(default=5, ge=1)
    min_negative_controls: int = Field(default=5, ge=1)
    require_representative: bool = True
    require_annotation_version: bool = True
    chromosome_level_policy: ChromosomePolicy = "warn"


class BenchmarkManifest(StrictBenchmarkModel):
    """Top-level biological benchmark manifest."""

    schema_version: Literal["1.0"] = "1.0"
    project: BenchmarkProject = Field(default_factory=BenchmarkProject)
    family: BenchmarkFamily = Field(default_factory=BenchmarkFamily)
    panel: BenchmarkPanel = Field(default_factory=BenchmarkPanel)
    acceptance: BenchmarkAcceptance = Field(default_factory=BenchmarkAcceptance)


class BenchmarkSpecies(StrictBenchmarkModel):
    """One row from the benchmark assembled-genome panel table."""

    species_id: BenchmarkId
    species_name: str
    group: str | None = None
    include: bool = True
    representative: bool = False
    input_kind: InputKind = "assembled_genome"
    assembly_accession: str | None = None
    assembly_level: AssemblyLevel = "unknown"
    annotation_version: str | None = None
    coordinate_system: str | None = None
    genome: Path | None = None
    genome_sha256: str | None = None
    gff3: Path | None = None
    gff3_sha256: str | None = None
    protein: Path | None = None
    protein_sha256: str | None = None
    cds: Path | None = None
    cds_sha256: str | None = None
    source_uri: str | None = None
    outgroup_species_id: BenchmarkId | None = None
    notes: str | None = None

    @model_validator(mode="before")
    @classmethod
    def normalize_tsv_values(cls, value: Any) -> Any:
        if not isinstance(value, dict):
            return value
        normalized: dict[str, Any] = {}
        for raw_key, raw_value in value.items():
            key = str(raw_key).strip()
            if not key:
                continue
            if isinstance(raw_value, str):
                text = raw_value.strip()
                if text == "":
                    continue
                if key in {"include", "representative"}:
                    lowered = text.lower()
                    if lowered in {"1", "true", "yes", "y"}:
                        normalized[key] = True
                        continue
                    if lowered in {"0", "false", "no", "n"}:
                        normalized[key] = False
                        continue
                normalized[key] = text
            elif raw_value is not None:
                normalized[key] = raw_value
        return normalized

    @model_validator(mode="after")
    def validate_species_id(self) -> BenchmarkSpecies:
        if "__" in self.species_id:
            raise ValueError("species_id cannot contain '__'; it is reserved for stable IDs")
        for field_name in (
            "genome_sha256",
            "gff3_sha256",
            "protein_sha256",
            "cds_sha256",
        ):
            checksum = getattr(self, field_name)
            if checksum is not None and not _SHA256_RE.fullmatch(checksum):
                raise ValueError(f"{field_name} must contain exactly 64 hexadecimal characters")
        return self


@dataclass(frozen=True, slots=True)
class BenchmarkCheck:
    """One audit gate with evidence and a minimal remediation action."""

    check_id: str
    scope: str
    item_id: str
    severity: CheckSeverity
    status: CheckStatus
    message: str
    evidence: str = ""
    remediation: str = ""


@dataclass(frozen=True, slots=True)
class InputFileRecord:
    """Resolved file state and observed checksum for one required input."""

    species_id: str
    field: str
    configured_path: str
    exists: bool
    size_bytes: int | None
    expected_sha256: str
    observed_sha256: str
    checksum_match: bool | None
    format_state: str


@dataclass(frozen=True, slots=True)
class BenchmarkAudit:
    """Complete in-memory readiness audit."""

    manifest_path: Path
    manifest_sha256: str
    species_table_path: Path
    species_table_sha256: str | None
    generated_at_utc: str
    project_name: str
    overall_status: OverallStatus
    checks: tuple[BenchmarkCheck, ...]
    species: tuple[BenchmarkSpecies, ...]
    files: tuple[InputFileRecord, ...]

    @property
    def blocking_failures(self) -> int:
        return sum(check.severity == "BLOCK" and check.status == "FAIL" for check in self.checks)

    @property
    def warnings(self) -> int:
        return sum(check.status == "WARN" for check in self.checks)

    @property
    def passed(self) -> int:
        return sum(check.status == "PASS" for check in self.checks)


def load_benchmark_manifest(path: Path) -> BenchmarkManifest:
    """Load and validate a benchmark YAML manifest."""

    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(resolved)
    data = yaml.safe_load(resolved.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Benchmark manifest root must be a YAML mapping")
    return BenchmarkManifest.model_validate(data)


def _resolve(path: Path | None, root: Path) -> Path | None:
    if path is None:
        return None
    expanded = path.expanduser()
    return expanded.resolve() if expanded.is_absolute() else (root / expanded).resolve()


def _sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _open_text(path: Path):  # type: ignore[no-untyped-def]
    if path.suffix.lower() == ".gz":
        return gzip.open(path, "rt", encoding="utf-8", errors="replace")
    return path.open("rt", encoding="utf-8", errors="replace")


def _format_state(path: Path, field: str) -> str:
    try:
        with _open_text(path) as handle:
            if field in {"genome", "protein", "cds"}:
                for line in handle:
                    stripped = line.lstrip("\ufeff").strip()
                    if not stripped:
                        continue
                    return "PASS" if stripped.startswith(">") else "FAIL_NOT_FASTA"
                return "FAIL_EMPTY_TEXT"
            if field == "gff3":
                saw_version = False
                for line in handle:
                    stripped = line.lstrip("\ufeff").rstrip("\n")
                    if not stripped:
                        continue
                    if stripped.startswith("##gff-version"):
                        saw_version = True
                        continue
                    if stripped.startswith("#"):
                        continue
                    columns = stripped.split("\t")
                    if len(columns) >= 9:
                        return "PASS_GFF3" if saw_version else "PASS_9_COLUMN_GFF"
                    return "FAIL_NOT_9_COLUMN_GFF"
                return "FAIL_NO_FEATURE_ROWS"
    except (OSError, UnicodeError) as error:
        return f"FAIL_READ:{type(error).__name__}"
    return "NOT_CHECKED"


def _check(
    checks: list[BenchmarkCheck],
    check_id: str,
    scope: str,
    item_id: str,
    severity: CheckSeverity,
    status: CheckStatus,
    message: str,
    evidence: str = "",
    remediation: str = "",
) -> None:
    checks.append(
        BenchmarkCheck(
            check_id=check_id,
            scope=scope,
            item_id=item_id,
            severity=severity,
            status=status,
            message=message,
            evidence=evidence,
            remediation=remediation,
        )
    )


def _read_species_table(
    path: Path,
    checks: list[BenchmarkCheck],
) -> list[BenchmarkSpecies]:
    records: list[BenchmarkSpecies] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            if reader.fieldnames is None:
                _check(
                    checks,
                    "BMG030",
                    "panel",
                    "species_table",
                    "BLOCK",
                    "FAIL",
                    "物种表缺少表头。",
                    str(path),
                    "使用 panfamflow benchmark init 生成标准 species.tsv。",
                )
                return records
            for row_number, row in enumerate(reader, start=2):
                try:
                    records.append(BenchmarkSpecies.model_validate(row))
                except ValidationError as error:
                    _check(
                        checks,
                        "BMG031",
                        "species_row",
                        f"row_{row_number}",
                        "BLOCK",
                        "FAIL",
                        "物种表行无法通过严格模式校验。",
                        str(error),
                        "修正该行字段、布尔值、路径或 SHA256 后重新审计。",
                    )
    except OSError as error:
        _check(
            checks,
            "BMG030",
            "panel",
            "species_table",
            "BLOCK",
            "FAIL",
            "无法读取物种表。",
            f"{path}: {error}",
            "恢复 species.tsv 或修正 benchmark.yaml 中的路径。",
        )
    return records


def _audit_truth_set(
    path: Path,
    manifest: BenchmarkManifest,
    known_species: set[str],
    checks: list[BenchmarkCheck],
) -> None:
    if not manifest.acceptance.require_manual_truth_set:
        _check(
            checks,
            "BMG060",
            "family",
            "manual_truth_set",
            "INFO",
            "INFO",
            "当前验收配置未要求人工正负例基线。",
        )
        return
    if not path.is_file() or path.stat().st_size == 0:
        _check(
            checks,
            "BMG060",
            "family",
            "manual_truth_set",
            "BLOCK",
            "FAIL",
            "人工正负例基线文件不存在或为空。",
            str(path),
            "填写 manual_review/manual_truth_set.tsv，并保留证据、审阅者和状态。",
        )
        return

    positive = 0
    negative = 0
    unknown_species: set[str] = set()
    invalid_statuses: set[str] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            required = {"species_id", "gene_id", "expected_status", "evidence", "reviewer"}
            fields = set(reader.fieldnames or [])
            missing = sorted(required - fields)
            if missing:
                _check(
                    checks,
                    "BMG061",
                    "family",
                    "manual_truth_set",
                    "BLOCK",
                    "FAIL",
                    "人工基线表缺少必需列。",
                    ", ".join(missing),
                    "按模板恢复 species_id、gene_id、expected_status、evidence、reviewer。",
                )
                return
            for row in reader:
                species_id = (row.get("species_id") or "").strip()
                gene_id = (row.get("gene_id") or "").strip()
                expected_status = (row.get("expected_status") or "").strip().upper()
                if not species_id or not gene_id or gene_id.upper().startswith("TODO"):
                    continue
                if species_id not in known_species:
                    unknown_species.add(species_id)
                if expected_status == "POSITIVE":
                    positive += 1
                elif expected_status == "NEGATIVE":
                    negative += 1
                elif expected_status not in {"UNCERTAIN", "NOT_ASSESSABLE"}:
                    invalid_statuses.add(expected_status or "<EMPTY>")
    except OSError as error:
        _check(
            checks,
            "BMG060",
            "family",
            "manual_truth_set",
            "BLOCK",
            "FAIL",
            "无法读取人工基线表。",
            f"{path}: {error}",
            "修复文件权限或编码后重新审计。",
        )
        return

    if invalid_statuses:
        _check(
            checks,
            "BMG062",
            "family",
            "manual_truth_set",
            "BLOCK",
            "FAIL",
            "人工基线包含非法 expected_status。",
            ", ".join(sorted(invalid_statuses)),
            "仅使用 POSITIVE、NEGATIVE、UNCERTAIN、NOT_ASSESSABLE。",
        )
    else:
        _check(
            checks,
            "BMG062",
            "family",
            "manual_truth_set",
            "BLOCK",
            "PASS",
            "人工基线状态枚举有效。",
        )

    if unknown_species:
        _check(
            checks,
            "BMG063",
            "family",
            "manual_truth_set",
            "BLOCK",
            "FAIL",
            "人工基线引用了物种表之外的 species_id。",
            ", ".join(sorted(unknown_species)),
            "统一 species_id，避免跨表主键漂移。",
        )
    else:
        _check(
            checks,
            "BMG063",
            "family",
            "manual_truth_set",
            "BLOCK",
            "PASS",
            "人工基线 species_id 均可映射到物种表。",
        )

    for check_id, label, observed, required_count in (
        (
            "BMG064",
            "阳性对照",
            positive,
            manifest.acceptance.min_positive_controls,
        ),
        (
            "BMG065",
            "阴性对照",
            negative,
            manifest.acceptance.min_negative_controls,
        ),
    ):
        control_status: CheckStatus = "PASS" if observed >= required_count else "FAIL"
        _check(
            checks,
            check_id,
            "family",
            "manual_truth_set",
            "BLOCK",
            control_status,
            f"{label}数量：{observed}；预注册最低数量：{required_count}。",
            remediation=(
                "补充可追溯的人工正负例，并在盲审前冻结。" if control_status == "FAIL" else ""
            ),
        )


def audit_benchmark(manifest_path: Path) -> BenchmarkAudit:
    """Audit a biological benchmark manifest without modifying source inputs."""

    resolved_manifest = manifest_path.expanduser().resolve()
    manifest = load_benchmark_manifest(resolved_manifest)
    root = resolved_manifest.parent
    checks: list[BenchmarkCheck] = []
    file_records: list[InputFileRecord] = []

    _check(
        checks,
        "BMG001",
        "manifest",
        "schema",
        "BLOCK",
        "PASS",
        "benchmark.yaml 已通过严格 Pydantic 模式校验。",
        f"schema_version={manifest.schema_version}",
    )

    acceptance_status: CheckStatus = (
        "PASS" if manifest.acceptance.approval_state == "approved" else "FAIL"
    )
    _check(
        checks,
        "BMG010",
        "acceptance",
        "approval_state",
        "BLOCK",
        acceptance_status,
        f"验收阈值冻结状态：{manifest.acceptance.approval_state}。",
        remediation=(
            "在查看最终组间差异前审阅阈值，并将 approval_state 改为 approved。"
            if acceptance_status == "FAIL"
            else ""
        ),
    )

    family_name_valid = manifest.family.name.strip().upper() not in _PLACEHOLDER_NAMES
    family_approved = manifest.family.approval_state == "approved"
    family_status: CheckStatus = (
        "PASS"
        if family_name_valid
        and (family_approved or not manifest.acceptance.require_family_approved)
        else "FAIL"
    )
    _check(
        checks,
        "BMG020",
        "family",
        "definition",
        "BLOCK",
        family_status,
        (f"目标家族={manifest.family.name!r}；冻结状态={manifest.family.approval_state}。"),
        remediation=(
            "明确家族名称、证据边界和人工审阅负责人，并将 family.approval_state 冻结为 approved。"
            if family_status == "FAIL"
            else ""
        ),
    )

    domain_ids = [*manifest.family.pfam_ids, *manifest.family.interpro_ids]
    domain_status: CheckStatus = "PASS" if domain_ids else "FAIL"
    _check(
        checks,
        "BMG021",
        "family",
        "domain_ids",
        "BLOCK",
        domain_status,
        "已配置家族 domain 标识。" if domain_ids else "未配置 Pfam/InterPro domain 标识。",
        ", ".join(domain_ids),
        "冻结至少一个经核验的 Pfam 或 InterPro ID。" if domain_status == "FAIL" else "",
    )

    evidence_paths = {
        "hmm": _resolve(manifest.family.hmm, root),
        "reference_proteins": _resolve(manifest.family.reference_proteins, root),
    }
    usable_evidence = [
        f"{name}={path}"
        for name, path in evidence_paths.items()
        if path is not None and path.is_file() and path.stat().st_size > 0
    ]
    evidence_status: CheckStatus = "PASS" if usable_evidence else "FAIL"
    _check(
        checks,
        "BMG022",
        "family",
        "search_evidence",
        "BLOCK",
        evidence_status,
        ("家族搜索证据文件可用。" if usable_evidence else "HMM 和参考蛋白均未提供可读的非空文件。"),
        "; ".join(usable_evidence),
        "提供版本化 HMM 和/或人工审阅的参考蛋白 FASTA。" if evidence_status == "FAIL" else "",
    )

    species_table = _resolve(manifest.panel.species_table, root)
    assert species_table is not None
    species_table_sha256: str | None = None
    species_records: list[BenchmarkSpecies] = []
    if not species_table.is_file() or species_table.stat().st_size == 0:
        _check(
            checks,
            "BMG030",
            "panel",
            "species_table",
            "BLOCK",
            "FAIL",
            "物种表不存在或为空。",
            str(species_table),
            "使用 benchmark init 模板建立 5–10 个 assembled-genome 条目。",
        )
    else:
        species_table_sha256 = _sha256_file(species_table)
        _check(
            checks,
            "BMG030",
            "panel",
            "species_table",
            "BLOCK",
            "PASS",
            "物种表存在且非空。",
            f"sha256={species_table_sha256}",
        )
        species_records = _read_species_table(species_table, checks)

    ids = [record.species_id for record in species_records]
    duplicates = sorted({species_id for species_id in ids if ids.count(species_id) > 1})
    duplicate_status: CheckStatus = "FAIL" if duplicates else "PASS"
    _check(
        checks,
        "BMG032",
        "panel",
        "species_id",
        "BLOCK",
        duplicate_status,
        "species_id 全局唯一。" if not duplicates else "物种表包含重复 species_id。",
        ", ".join(duplicates),
        "修正重复主键后重新审计。" if duplicates else "",
    )

    included = [record for record in species_records if record.include]
    assembled = [record for record in included if record.input_kind == "assembled_genome"]
    panel_size_ok = manifest.panel.min_genomes <= len(
        assembled
    ) <= manifest.panel.max_genomes and len(assembled) == len(included)
    _check(
        checks,
        "BMG040",
        "panel",
        "panel_size",
        "BLOCK",
        "PASS" if panel_size_ok else "FAIL",
        (
            f"纳入条目={len(included)}；assembled genome={len(assembled)}；"
            f"要求={manifest.panel.min_genomes}–{manifest.panel.max_genomes}。"
        ),
        remediation=(
            "仅纳入相互独立、已组装并注释的基因组；同一参考坐标的 BAM/VCF 不能作为多个 genome。"
            if not panel_size_ok
            else ""
        ),
    )

    representatives = [record.species_id for record in included if record.representative]
    if manifest.acceptance.require_representative:
        representative_ok = len(representatives) == 1
        _check(
            checks,
            "BMG041",
            "panel",
            "representative",
            "BLOCK",
            "PASS" if representative_ok else "FAIL",
            f"代表基因组数量={len(representatives)}。",
            ", ".join(representatives),
            "指定且仅指定一个代表基因组。" if not representative_ok else "",
        )

    groups = sorted({record.group for record in included if record.group})
    _check(
        checks,
        "BMG042",
        "panel",
        "group_diversity",
        "WARN",
        "PASS" if len(groups) >= 2 else "WARN",
        f"已定义群体数={len(groups)}。",
        ", ".join(groups),
        "如需群体比较，至少预先定义两个有生物学依据的群体。" if len(groups) < 2 else "",
    )

    known_species = set(ids)
    for record in included:
        item = record.species_id
        if record.input_kind != "assembled_genome":
            _check(
                checks,
                "BMG100",
                "species",
                item,
                "BLOCK",
                "FAIL",
                "纳入项不是独立 assembled genome。",
                f"input_kind={record.input_kind}",
                "将 reference-aligned BAM/VCF 样本移出 genome panel。",
            )
        else:
            _check(
                checks,
                "BMG100",
                "species",
                item,
                "BLOCK",
                "PASS",
                "输入类型为 assembled_genome。",
            )

        required_metadata: dict[str, str | None] = {
            "assembly_accession": record.assembly_accession,
            "coordinate_system": record.coordinate_system,
        }
        if manifest.acceptance.require_annotation_version:
            required_metadata["annotation_version"] = record.annotation_version
        missing_metadata = sorted(name for name, value in required_metadata.items() if not value)
        _check(
            checks,
            "BMG101",
            "species",
            item,
            "BLOCK",
            "PASS" if not missing_metadata else "FAIL",
            "版本与坐标元数据完整。" if not missing_metadata else "版本或坐标元数据缺失。",
            ", ".join(missing_metadata),
            "补充 assembly accession、annotation version 和 coordinate system。"
            if missing_metadata
            else "",
        )

        if record.assembly_level != "chromosome":
            policy = manifest.acceptance.chromosome_level_policy
            if policy != "ignore":
                _check(
                    checks,
                    "BMG102",
                    "species",
                    item,
                    "BLOCK" if policy == "block" else "WARN",
                    "FAIL" if policy == "block" else "WARN",
                    f"assembly_level={record.assembly_level}，染色体定位与共线性解释将受限。",
                    remediation="优先替换为 chromosome-level assembly，或在验收中明确降级。",
                )
        else:
            _check(
                checks,
                "BMG102",
                "species",
                item,
                "WARN",
                "PASS",
                "assembly_level=chromosome。",
            )

        if (
            record.outgroup_species_id is not None
            and record.outgroup_species_id not in known_species
        ):
            _check(
                checks,
                "BMG103",
                "species",
                item,
                "BLOCK",
                "FAIL",
                "outgroup_species_id 未在物种表中定义。",
                record.outgroup_species_id,
                "加入外群条目或修正外群主键。",
            )

        for field in _REQUIRED_SPECIES_PATHS:
            configured = getattr(record, field)
            resolved = _resolve(configured, root)
            expected = getattr(record, f"{field}_sha256") or ""
            exists = resolved is not None and resolved.is_file()
            size = resolved.stat().st_size if exists and resolved is not None else None
            observed = _sha256_file(resolved) if exists and resolved is not None else ""
            checksum_match: bool | None = None
            if expected and observed:
                checksum_match = expected.lower() == observed.lower()
            format_state = (
                _format_state(resolved, field) if exists and resolved is not None else "MISSING"
            )
            file_records.append(
                InputFileRecord(
                    species_id=item,
                    field=field,
                    configured_path=str(configured or ""),
                    exists=exists,
                    size_bytes=size,
                    expected_sha256=expected,
                    observed_sha256=observed,
                    checksum_match=checksum_match,
                    format_state=format_state,
                )
            )
            _check(
                checks,
                "BMG110",
                f"species.{field}",
                item,
                "BLOCK",
                "PASS" if exists and size and size > 0 else "FAIL",
                f"{field} 文件存在且非空。"
                if exists and size and size > 0
                else f"{field} 文件缺失或为空。",
                str(resolved) if resolved is not None else "<not configured>",
                f"提供与同一 assembly/annotation 版本配套的 {field} 文件。"
                if not (exists and size and size > 0)
                else "",
            )

            if manifest.acceptance.require_sha256:
                _check(
                    checks,
                    "BMG111",
                    f"species.{field}",
                    item,
                    "BLOCK",
                    "PASS" if expected else "FAIL",
                    f"{field} 已登记预期 SHA256。" if expected else f"{field} 未登记预期 SHA256。",
                    expected,
                    "计算源文件 SHA256 并在 species.tsv 中冻结。" if not expected else "",
                )
                if expected and observed:
                    _check(
                        checks,
                        "BMG112",
                        f"species.{field}",
                        item,
                        "BLOCK",
                        "PASS" if checksum_match else "FAIL",
                        f"{field} SHA256 匹配。" if checksum_match else f"{field} SHA256 不匹配。",
                        f"expected={expected}; observed={observed}",
                        "确认文件版本，禁止用同名不同内容替换输入。" if not checksum_match else "",
                    )

            if exists:
                format_ok = format_state.startswith("PASS")
                _check(
                    checks,
                    "BMG113",
                    f"species.{field}",
                    item,
                    "BLOCK",
                    "PASS" if format_ok else "FAIL",
                    f"{field} 轻量格式嗅探：{format_state}。",
                    remediation="确认 FASTA/GFF3 编码与文件类型。" if not format_ok else "",
                )

    truth_path = _resolve(manifest.family.manual_truth_set, root)
    assert truth_path is not None
    _audit_truth_set(truth_path, manifest, known_species, checks)

    blocking_failures = sum(
        check.severity == "BLOCK" and check.status == "FAIL" for check in checks
    )
    warnings = sum(check.status == "WARN" for check in checks)
    overall: OverallStatus
    if blocking_failures:
        overall = "BLOCKED"
    elif warnings:
        overall = "REVIEW"
    else:
        overall = "READY"

    return BenchmarkAudit(
        manifest_path=resolved_manifest,
        manifest_sha256=_sha256_file(resolved_manifest),
        species_table_path=species_table,
        species_table_sha256=species_table_sha256,
        generated_at_utc=datetime.now(UTC).isoformat(),
        project_name=manifest.project.name,
        overall_status=overall,
        checks=tuple(checks),
        species=tuple(species_records),
        files=tuple(file_records),
    )


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.partial.{os.getpid()}")
    temporary.write_text(text, encoding="utf-8")
    os.replace(temporary, path)


def _write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.partial.{os.getpid()}")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)
    os.replace(temporary, path)


def _species_dict(record: BenchmarkSpecies) -> dict[str, Any]:
    return record.model_dump(mode="json", exclude_none=False)


def _render_markdown(audit: BenchmarkAudit) -> str:
    lines = [
        f"# {audit.project_name} 生物学验收门",
        "",
        f"- 总体状态：**{audit.overall_status}**",
        f"- 阻断项：{audit.blocking_failures}",
        f"- 警告项：{audit.warnings}",
        f"- 通过项：{audit.passed}",
        f"- 生成时间（UTC）：`{audit.generated_at_utc}`",
        f"- Manifest SHA256：`{audit.manifest_sha256}`",
        "",
        "| Check | Scope | Item | Severity | Status | Message | Evidence | Remediation |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for check in audit.checks:
        cells = [
            check.check_id,
            check.scope,
            check.item_id,
            check.severity,
            check.status,
            check.message,
            check.evidence,
            check.remediation,
        ]
        escaped = [str(cell).replace("|", "\\|").replace("\n", "<br>") for cell in cells]
        lines.append("| " + " | ".join(escaped) + " |")
    lines.append("")
    return "\n".join(lines)


def _render_html(audit: BenchmarkAudit) -> str:
    status_class = audit.overall_status.lower()
    rows = []
    for check in audit.checks:
        rows.append(
            "<tr>"
            f"<td><code>{html.escape(check.check_id)}</code></td>"
            f"<td>{html.escape(check.scope)}</td>"
            f"<td>{html.escape(check.item_id)}</td>"
            f"<td><span class='sev {html.escape(check.severity.lower())}'>{html.escape(check.severity)}</span></td>"
            f"<td><span class='check {html.escape(check.status.lower())}'>{html.escape(check.status)}</span></td>"
            f"<td>{html.escape(check.message)}</td>"
            f"<td><code>{html.escape(check.evidence)}</code></td>"
            f"<td>{html.escape(check.remediation)}</td>"
            "</tr>"
        )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(audit.project_name)} 生物学验收门</title>
<style>
:root {{ color-scheme: light; font-family: Arial, "Noto Sans CJK SC", sans-serif; }}
body {{ margin: 0; background: #f6f7f9; color: #1f2937; }}
main {{ max-width: 1500px; margin: 0 auto; padding: 28px; }}
h1 {{ margin: 0 0 8px; }}
.meta {{ color: #4b5563; margin-bottom: 22px; }}
.cards {{ display: grid; grid-template-columns: repeat(4, minmax(160px, 1fr)); gap: 12px; margin: 18px 0 24px; }}
.card {{ background: white; border: 1px solid #d1d5db; border-radius: 10px; padding: 16px; }}
.card strong {{ display: block; font-size: 26px; margin-top: 4px; }}
.overall {{ border-left: 8px solid #9ca3af; }}
.overall.ready {{ border-left-color: #238636; }}
.overall.review {{ border-left-color: #b7791f; }}
.overall.blocked {{ border-left-color: #b42318; }}
table {{ width: 100%; border-collapse: collapse; background: white; font-size: 13px; }}
th, td {{ border: 1px solid #d1d5db; padding: 8px; text-align: left; vertical-align: top; }}
th {{ background: #eef2f7; position: sticky; top: 0; }}
code {{ white-space: pre-wrap; overflow-wrap: anywhere; }}
.sev, .check {{ display: inline-block; border-radius: 999px; padding: 2px 8px; font-weight: 700; }}
.sev.block, .check.fail {{ background: #fee4e2; color: #912018; }}
.sev.warn, .check.warn {{ background: #fff4cc; color: #7a4d00; }}
.sev.info, .check.info {{ background: #e6f0ff; color: #174ea6; }}
.check.pass {{ background: #dcfce7; color: #166534; }}
.note {{ background: #fff; border: 1px solid #d1d5db; border-radius: 10px; padding: 14px 16px; margin: 0 0 18px; }}
@media (max-width: 900px) {{ .cards {{ grid-template-columns: 1fr 1fr; }} main {{ padding: 14px; }} }}
</style>
</head>
<body>
<main>
<h1>PanFamFlow 生物学验收门</h1>
<div class="meta">项目：{html.escape(audit.project_name)}；生成时间（UTC）：{html.escape(audit.generated_at_utc)}</div>
<div class="note">该页面只判断真实生物学 benchmark 的输入和预注册状态是否满足启动条件。软件 CI 通过不等于生物学验收通过。</div>
<section class="cards">
<div class="card overall {status_class}">总体状态<strong>{html.escape(audit.overall_status)}</strong></div>
<div class="card">阻断项<strong>{audit.blocking_failures}</strong></div>
<div class="card">警告项<strong>{audit.warnings}</strong></div>
<div class="card">通过项<strong>{audit.passed}</strong></div>
</section>
<table>
<thead><tr><th>检查项</th><th>范围</th><th>对象</th><th>级别</th><th>状态</th><th>说明</th><th>证据</th><th>最小修复动作</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>
</main>
</body>
</html>
"""


def _write_workbook(audit: BenchmarkAudit, path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = "summary"
    summary_rows = [
        ("project_name", audit.project_name),
        ("overall_status", audit.overall_status),
        ("blocking_failures", audit.blocking_failures),
        ("warnings", audit.warnings),
        ("passed", audit.passed),
        ("generated_at_utc", audit.generated_at_utc),
        ("manifest_path", str(audit.manifest_path)),
        ("manifest_sha256", audit.manifest_sha256),
        ("species_table_path", str(audit.species_table_path)),
        ("species_table_sha256", audit.species_table_sha256 or ""),
    ]
    for summary_row in summary_rows:
        summary.append(summary_row)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 80

    checks_sheet = workbook.create_sheet("checks")
    check_fields = list(BenchmarkCheck.__dataclass_fields__)
    checks_sheet.append(check_fields)
    for check in audit.checks:
        checks_sheet.append([getattr(check, field) for field in check_fields])
    files_sheet = workbook.create_sheet("input_files")
    file_fields = list(InputFileRecord.__dataclass_fields__)
    files_sheet.append(file_fields)
    for record in audit.files:
        files_sheet.append([getattr(record, field) for field in file_fields])
    species_sheet = workbook.create_sheet("species")
    species_rows = [_species_dict(record) for record in audit.species]
    species_fields = sorted({key for row in species_rows for key in row})
    species_sheet.append(species_fields)
    for species_row in species_rows:
        species_sheet.append([species_row.get(field, "") for field in species_fields])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.partial.{os.getpid()}")
    workbook.save(temporary)
    os.replace(temporary, path)


def write_benchmark_audit(audit: BenchmarkAudit, output_dir: Path) -> dict[str, Path]:
    """Write non-destructive TSV/XLSX/JSON/Markdown/HTML audit outputs."""

    target = output_dir.expanduser().resolve()
    if target.exists() and any(target.iterdir()):
        raise FileExistsError(f"Refusing to overwrite non-empty audit directory: {target}")
    target.mkdir(parents=True, exist_ok=True)

    check_rows = [asdict(check) for check in audit.checks]
    file_rows = [asdict(record) for record in audit.files]
    species_rows = [_species_dict(record) for record in audit.species]
    check_fields = list(BenchmarkCheck.__dataclass_fields__)
    file_fields = list(InputFileRecord.__dataclass_fields__)
    species_fields = sorted({key for row in species_rows for key in row})

    paths = {
        "checks_tsv": target / "benchmark_readiness.tsv",
        "files_tsv": target / "input_files.tsv",
        "species_tsv": target / "species_snapshot.tsv",
        "xlsx": target / "benchmark_readiness.xlsx",
        "json": target / "benchmark_readiness.json",
        "markdown": target / "benchmark_readiness.md",
        "html": target / "benchmark_readiness.html",
        "sha256": target / "SHA256SUMS.tsv",
    }
    _write_tsv(paths["checks_tsv"], check_fields, check_rows)
    _write_tsv(paths["files_tsv"], file_fields, file_rows)
    _write_tsv(paths["species_tsv"], species_fields, species_rows)
    _write_workbook(audit, paths["xlsx"])
    payload = {
        "schema_version": "1.0",
        "project_name": audit.project_name,
        "overall_status": audit.overall_status,
        "blocking_failures": audit.blocking_failures,
        "warnings": audit.warnings,
        "passed": audit.passed,
        "generated_at_utc": audit.generated_at_utc,
        "manifest_path": str(audit.manifest_path),
        "manifest_sha256": audit.manifest_sha256,
        "species_table_path": str(audit.species_table_path),
        "species_table_sha256": audit.species_table_sha256,
        "checks": check_rows,
        "species": species_rows,
        "input_files": file_rows,
    }
    _atomic_write_text(paths["json"], json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    _atomic_write_text(paths["markdown"], _render_markdown(audit))
    _atomic_write_text(paths["html"], _render_html(audit))

    checksum_rows = []
    for name, path in paths.items():
        if name == "sha256":
            continue
        checksum_rows.append(
            {
                "file": path.name,
                "size_bytes": path.stat().st_size,
                "sha256": _sha256_file(path),
            }
        )
    _write_tsv(paths["sha256"], ["file", "size_bytes", "sha256"], checksum_rows)
    return paths


def default_audit_output_dir(manifest_path: Path) -> Path:
    """Return a new timestamped audit directory beside the manifest."""

    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%S.%fZ")
    return manifest_path.expanduser().resolve().parent / "audits" / stamp


def initialize_benchmark(destination: Path) -> tuple[Path, ...]:
    """Create a non-destructive benchmark intake skeleton."""

    target = destination.expanduser().resolve()
    if target.exists() and not target.is_dir():
        raise NotADirectoryError(target)
    if target.exists() and any(target.iterdir()):
        raise FileExistsError(f"Refusing to overwrite non-empty benchmark directory: {target}")
    target.mkdir(parents=True, exist_ok=True)
    template_root = resources.files("panfamflow.templates").joinpath("benchmark")
    relative_files = (
        Path("benchmark.yaml"),
        Path("species.tsv"),
        Path("decision_log.tsv"),
        Path("README.zh-CN.md"),
        Path("manual_review/manual_truth_set.tsv"),
    )
    written: list[Path] = []
    for relative in relative_files:
        source = template_root.joinpath(*relative.parts)
        destination_path = target / relative
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        destination_path.write_text(source.read_text(encoding="utf-8"), encoding="utf-8")
        written.append(destination_path)
    schema_path = target / ".panfamflow" / "benchmark.schema.json"
    schema_path.parent.mkdir(parents=True, exist_ok=True)
    schema_path.write_text(
        json.dumps(BenchmarkManifest.model_json_schema(), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    written.append(schema_path)
    gitignore_path = target / ".gitignore"
    gitignore_path.write_text(
        "inputs/\naudits/\n*.partial.*\n.DS_Store\n",
        encoding="utf-8",
    )
    written.append(gitignore_path)
    for directory in ("inputs", "references", "audits"):
        (target / directory).mkdir(exist_ok=True)
    return tuple(written)
